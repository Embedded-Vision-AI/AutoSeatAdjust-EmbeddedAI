import os
import pandas as pd
import xgboost as xgb
import joblib
from openpyxl import load_workbook
from DTreeClassifier import (
    DATA_XLSX_PATH,
    MODEL_PATH,
    FEATURE_COL,
    TARGET_COLS,
    load_data,
    train_models,
    save_models
)

#utlilities
def load_models(model_path: str):
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"No model file at {model_path}. Run training first.")
    return joblib.load(model_path)

def prompt_float(msg):
    while True:
        try:
            return float(input(msg))
        except ValueError:
            print("Please enter a numeric value.")

#appending new data values from terminal to the dataset
def append_row_to_excel(path, row_dict):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        new_values = [row_dict.get(h) for h in headers]
        ws.append(new_values)
        wb.save(path)
    else:
        #if dataset doesn't exit, making a new one
        df = pd.DataFrame([row_dict])
        df.to_excel(path, index=False, engine="openpyxl")
    print("[INFO] Row appended to Excel.")

if __name__ == "__main__":
    if not os.path.exists(MODEL_PATH):
        print("[INFO] No model found – training first.")
        X0, y0 = load_data(DATA_XLSX_PATH)
        save_models(train_models(X0, y0), MODEL_PATH)

    models = load_models(MODEL_PATH)

    #inference loop
    height = prompt_float("\nEnter driver Height (cm): ")

    X_infer = pd.DataFrame({FEATURE_COL: [height]})
    preds   = {t: models[t].predict(X_infer)[0] for t in TARGET_COLS}

    print("\n[RECOMMENDED SETTINGS]")
    for k, v in preds.items():
        print(f"  {k}: {v:.2f}")

    if input("\nAccept these settings? (y/n): ").strip().lower() == "y":
        print("Great! Exiting.")
        exit(0)

#collect corrected value
    corrected = {}
    for col in TARGET_COLS:
        corrected[col] = prompt_float(f"Enter desired {col}: ")

    #making new rows for feature + targets
    new_row = {FEATURE_COL: height}
    new_row.update(corrected)

    append_row_to_excel(DATA_XLSX_PATH, new_row)

#triggering retraining after dataset is updated with new datapoints
    print("[INFO] Retraining with new data…")
    X_all, y_all = load_data(DATA_XLSX_PATH)
    new_models   = train_models(X_all, y_all)
    save_models(new_models, MODEL_PATH)
    print("[DONE] Model retrained with new datapoint.")